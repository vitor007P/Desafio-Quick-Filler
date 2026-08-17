"""Geração das planilhas nos três formatos aceitos.

Os três saem da mesma `Grade`, então o que a tela mostra, o que o `.xlsx`
colore e o que o `.csv` lista são necessariamente o mesmo dado.
"""

from __future__ import annotations

import csv
import io
import json

from app.grid import AMARELO, BORDA_VERMELHA, CABECALHO_FUNDO, VERMELHO, Grade

FORMATOS = {
    "xlsx": ("xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
    "csv": ("csv", "text/csv; charset=utf-8"),
    "json": ("json", "application/json; charset=utf-8"),
}


def gerar(grade: Grade, formato: str) -> bytes:
    """Ponto de entrada único: escolhe o gerador certo pelo nome do formato."""
    if formato == "xlsx":
        return _xlsx(grade)
    if formato == "csv":
        return _csv(grade)
    if formato == "json":
        return _json(grade)
    raise ValueError(f"formato não suportado: {formato}")


def _xlsx(grade: Grade) -> bytes:
    """Monta a planilha Excel: cabeçalho colorido, linhas com aviso destacadas."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active  # a primeira (e única) aba da planilha
    ws.title = "Transcrição"

    fundo_cabecalho = PatternFill("solid", fgColor=CABECALHO_FUNDO)
    fonte_cabecalho = Font(bold=True, color="FFFFFF")
    preenchimentos = {
        "amarelo": PatternFill("solid", fgColor=AMARELO),
        "vermelho": PatternFill("solid", fgColor=VERMELHO),
    }
    borda_esquerda = Border(left=Side(style="medium", color=BORDA_VERMELHA))

    ws.append(grade.colunas)  # primeira linha da planilha = cabeçalho
    for celula in ws[1]:
        celula.fill = fundo_cabecalho
        celula.font = fonte_cabecalho
        celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for linha in grade.linhas:
        ws.append(linha.celulas)
        indice = ws.max_row  # linha que acabamos de inserir
        severidade = linha.severidade
        if severidade == "nenhum":
            continue  # linha sem aviso: nenhuma cor extra

        preenchimento = preenchimentos[severidade]
        for coluna in range(1, len(grade.colunas) + 1):
            ws.cell(row=indice, column=coluna).fill = preenchimento

        if severidade == "vermelho":
            ws.cell(row=indice, column=1).border = borda_esquerda

        # O motivo fica no comentário da primeira célula: a planilha continua
        # com as colunas exatas do contrato, e o revisor ainda vê o porquê.
        if linha.mensagens:
            ws.cell(row=indice, column=1).comment = _comentario("; ".join(linha.mensagens))

    for i, titulo in enumerate(grade.colunas, start=1):
        largura = max(len(titulo) + 2, 10)
        for linha in grade.linhas:
            if i <= len(linha.celulas):
                largura = max(largura, min(len(str(linha.celulas[i - 1])) + 2, 40))
        ws.column_dimensions[get_column_letter(i)].width = largura

    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _comentario(texto: str):
    """Cria uma anotação (comment) do Excel, do tamanho certo pra caber o texto."""
    from openpyxl.comments import Comment

    comentario = Comment(texto, "Transcrição")
    comentario.width = 300
    comentario.height = 60
    return comentario


def _csv(grade: Grade) -> bytes:
    """Monta o CSV: mesma tabela do Excel, sem cores (CSV não suporta estilo)."""
    buffer = io.StringIO()
    # Ponto e vírgula porque os valores monetários usam vírgula decimal; é o
    # que o Excel em pt-BR abre sem pedir importação.
    escritor = csv.writer(buffer, delimiter=";", lineterminator="\r\n")
    escritor.writerow(grade.colunas)
    for linha in grade.linhas:
        escritor.writerow(linha.celulas)
    # BOM para o Excel reconhecer UTF-8.
    return b"\xef\xbb\xbf" + buffer.getvalue().encode("utf-8")


def _json(grade: Grade) -> bytes:
    """Monta o JSON: a mesma grade, "crua", para quem quiser processar por script."""
    return json.dumps(grade.to_dict(), ensure_ascii=False, indent=2).encode("utf-8")


__all__ = ["FORMATOS", "gerar"]
